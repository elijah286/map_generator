"""
Shared map generation logic for Excel → geocoded SVG maps.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from typing import Callable, Dict, List, Optional, Set

import matplotlib.pyplot as plt
import pandas as pd
import cartopy.crs as ccrs
import cartopy.feature as cfeature
import openai

LogFn = Callable[[str], None]


def _noop_log(msg: str) -> None:
    pass


def load_cache(cache_path: str) -> dict:
    if os.path.exists(cache_path):
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache_path: str, cache: dict) -> None:
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)


def parse_lat_lon_from_gpt(response_text: str) -> Optional[tuple]:
    """Extract first valid lat,lon pair from GPT output (handles extra lines)."""
    text = response_text.strip()
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # Match "12.34, -56.78" or similar
        m = re.search(
            r"(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)",
            line,
        )
        if m:
            try:
                lat, lon = float(m.group(1)), float(m.group(2))
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    return lat, lon
            except ValueError:
                continue
    return None


def gpt_geocode(location_string: str, api_key: str, log: LogFn = _noop_log) -> Optional[tuple]:
    if not location_string or not api_key:
        if not api_key:
            log("OpenAI API key is missing. Set OPENAI_API_KEY or enter it in the app.")
        return None
    try:
        log(f"Geocoding: {location_string}")
        client = openai.OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a geolocation API. Only respond with valid latitude and longitude "
                        "in the format: 12.34, -56.78. Do not add any explanation or extra words."
                    ),
                },
                {
                    "role": "user",
                    "content": f"What are the latitude and longitude coordinates for {location_string}?",
                },
            ],
            temperature=0,
        )
        response_text = response.choices[0].message.content.strip()
        log(f"  → {response_text}")
        coords = parse_lat_lon_from_gpt(response_text)
        if coords:
            return coords[0], coords[1]
        log(f"  ✗ Could not parse coordinates from response.")
    except Exception as e:
        log(f"  ✗ API error: {e}")
    return None


def geocode_location(
    location_string: str,
    cache: dict,
    cache_path: str,
    api_key: str,
    log: LogFn = _noop_log,
) -> Optional[tuple]:
    key = location_string.strip()
    if not key:
        return None
    if key in cache:
        v = cache[key]
        log(f"Cached: {key}")
        return tuple(v) if isinstance(v, (list, tuple)) else None
    for k, v in cache.items():
        if k.lower() == key.lower():
            log(f"Cached: {key}")
            return tuple(v) if isinstance(v, (list, tuple)) else None

    coords = gpt_geocode(key, api_key, log=log)
    if coords:
        cache[key] = [coords[0], coords[1]]
        save_cache(cache_path, cache)
    else:
        log(f"  ✗ Failed: {key}")
    return coords


def list_sheet_names(path: str) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def resolve_excel_path(path: str) -> Optional[str]:
    if not os.path.exists(path):
        return None
    actual = path
    if os.path.islink(path):
        target = os.readlink(path)
        if not os.path.isabs(target):
            target = os.path.join(os.path.dirname(path), target)
        actual = target
    return actual if os.path.exists(actual) else None


def load_user_events_df(path: str, sheet: str, only_on_map: bool) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)

    def construct_location(row):
        parts = [
            str(row.get(col, "")).strip()
            for col in ["City", "State", "Country"]
            if pd.notna(row.get(col, "")) and str(row.get(col, "")).strip()
        ]
        return ", ".join(parts)

    df = df.copy()
    df["Location String"] = df.apply(construct_location, axis=1)
    if only_on_map and "On Map?" in df.columns:
        df = df[df["On Map?"].astype(str).str.lower().str.strip() == "yes"]
    return df


def load_university_df(path: str, sheet: str, log: LogFn) -> Optional[pd.DataFrame]:
    actual = resolve_excel_path(path)
    if not actual:
        log(f"File not found: {path}")
        return None

    df = None
    for engine in ["openpyxl", "xlrd", None]:
        try:
            if engine:
                try:
                    df = pd.read_excel(actual, sheet_name=sheet, engine=engine)
                    start_row = 0
                    for idx, row in df.iterrows():
                        if str(row.iloc[0]).strip().lower() == "university":
                            start_row = idx
                            break
                    if start_row > 0:
                        df = pd.read_excel(
                            actual, sheet_name=sheet, engine=engine, header=start_row
                        )
                except Exception:
                    df = pd.read_excel(actual, sheet_name=sheet, engine=engine, header=3)
            else:
                df = pd.read_excel(actual, sheet_name=sheet, header=3)
            break
        except Exception as e:
            log(f"Read attempt failed ({engine}): {e}")
            continue

    if df is None:
        return None

    def construct_location(row):
        university = str(row.iloc[0] if pd.notna(row.iloc[0]) else "").strip()
        location = str(row.iloc[1] if pd.notna(row.iloc[1]) else "").strip()
        if university.lower() == "university":
            return ""
        if university and location:
            return f"{university}, {location}"
        if location:
            return location
        if university:
            return university
        return ""

    df = df.copy()
    df["Location String"] = df.apply(construct_location, axis=1)
    df = df[df["Location String"] != ""]
    return df


def _scatter_points(ax, data: pd.DataFrame, use_size: bool) -> None:
    for _, row in data.iterrows():
        lat, lon = row["Latitude"], row["Longitude"]
        size = 50
        if use_size and "Member Count" in data.columns and pd.notna(row.get("Member Count")):
            try:
                size = float(row["Member Count"]) * 2
            except (TypeError, ValueError):
                size = 50
        ax.scatter(
            lon,
            lat,
            s=size,
            color="#1a7f37",
            alpha=0.85,
            edgecolor="#0d3d1a",
            linewidth=0.5,
            zorder=5,
            transform=ccrs.PlateCarree(),
        )


def _style_axes(ax) -> None:
    ax.add_feature(cfeature.COASTLINE, linewidth=0.4)
    ax.add_feature(cfeature.BORDERS, linewidth=0.3, edgecolor="#999999")
    ax.add_feature(cfeature.OCEAN, facecolor="#e8eef5")
    ax.add_feature(cfeature.LAND, facecolor="#e4e2dc")


REGION_FILTERS: Dict[str, Callable[[pd.DataFrame], pd.DataFrame]] = {
    "world": lambda d: d,
    "europe": lambda d: d[
        (d["Latitude"] >= 35)
        & (d["Latitude"] <= 70)
        & (d["Longitude"] >= -10)
        & (d["Longitude"] <= 40)
    ],
    "asia": lambda d: d[
        (d["Latitude"] >= -10)
        & (d["Latitude"] <= 55)
        & (d["Longitude"] >= 60)
        & (d["Longitude"] <= 150)
    ],
    "north_america": lambda d: d[
        (d["Latitude"] >= 15)
        & (d["Latitude"] <= 75)
        & (d["Longitude"] >= -170)
        & (d["Longitude"] <= -50)
    ],
    "south_america": lambda d: d[
        (d["Latitude"] >= -60)
        & (d["Latitude"] <= 15)
        & (d["Longitude"] >= -85)
        & (d["Longitude"] <= -30)
    ],
}

REGION_EXTENTS = {
    "world": None,
    "europe": [-10, 40, 35, 70],
    "asia": [60, 150, -10, 55],
    "north_america": [-170, -50, 15, 75],
    "south_america": [-85, -30, -60, 15],
}

REGION_SUFFIX = {
    "world": "",
    "europe": "_europe",
    "asia": "_asia",
    "north_america": "_north_america",
    "south_america": "_south_america",
}


def plot_region(
    df: pd.DataFrame,
    region: str,
    use_size: bool,
    out_path: str,
) -> None:
    filt = REGION_FILTERS[region]
    extent = REGION_EXTENTS[region]
    data = filt(df)

    fig = plt.figure(figsize=(18, 9))
    ax = plt.axes(projection=ccrs.PlateCarree())
    _style_axes(ax)
    if extent is None:
        ax.set_global()
    else:
        ax.set_extent(extent, crs=ccrs.PlateCarree())
    _scatter_points(ax, data, use_size)
    plt.tight_layout()
    plt.savefig(out_path, format="svg", bbox_inches="tight")
    plt.close(fig)


@dataclass
class GenerationConfig:
    excel_path: str
    sheet_name: str
    mode: str  # "user_events" | "universities"
    only_on_map: bool = True
    use_member_count_size: bool = False
    output_dir: str = ""
    output_basename: str = "map"
    regions: Set[str] = field(
        default_factory=lambda: {"world", "europe", "asia", "north_america", "south_america"}
    )
    cache_path: str = ""
    api_key: str = ""
    open_outputs: bool = True


def default_cache_path() -> str:
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "location_cache.json")


def run_generation(cfg: GenerationConfig, log: LogFn = _noop_log) -> List[str]:
    api_key = cfg.api_key or os.environ.get("OPENAI_API_KEY", "")
    cache_path = cfg.cache_path or default_cache_path()
    out_dir = cfg.output_dir or os.path.dirname(os.path.abspath(__file__))
    os.makedirs(out_dir, exist_ok=True)
    base = cfg.output_basename.strip() or "map"

    cache = load_cache(cache_path)

    if cfg.mode == "user_events":
        df = load_user_events_df(cfg.excel_path, cfg.sheet_name, cfg.only_on_map)
    else:
        df = load_university_df(cfg.excel_path, cfg.sheet_name, log)
        if df is None or df.empty:
            log("No data loaded.")
            return []

    if df.empty:
        log("No rows to plot after filtering.")
        return []

    lats, lons = [], []
    for loc in df["Location String"]:
        c = geocode_location(loc, cache, cache_path, api_key, log=log)
        if c:
            lats.append(c[0])
            lons.append(c[1])
        else:
            lats.append(None)
            lons.append(None)

    df = df.copy()
    df["Latitude"] = lats
    df["Longitude"] = lons
    df = df.dropna(subset=["Latitude", "Longitude"])
    log(f"Plotted {len(df)} locations with coordinates.")

    written: List[str] = []
    for region in ("world", "europe", "asia", "north_america", "south_america"):
        if region not in cfg.regions:
            continue
        suf = REGION_SUFFIX[region]
        name = f"{base}{suf}.svg" if suf else f"{base}.svg"
        out_path = os.path.join(out_dir, name)
        plot_region(df, region, cfg.use_member_count_size, out_path)
        log(f"Saved: {out_path}")
        written.append(out_path)

    if cfg.open_outputs:
        sys_platform_open(written)
    return written


def sys_platform_open(paths: List[str]) -> bool:
    if not paths:
        return False
    if sys.platform == "darwin":
        for p in paths:
            subprocess.run(["open", p], check=False)
        return True
    if sys.platform == "win32":
        for p in paths:
            os.startfile(p)  # type: ignore[attr-defined]
        return True
    for p in paths:
        subprocess.run(["xdg-open", p], check=False)
    return True
